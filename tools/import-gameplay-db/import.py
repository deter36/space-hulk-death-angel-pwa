import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "docs" / "source" / "Death_Angel_Base_Game_Gameplay_Database.xlsx"
OUTPUT = ROOT / "src" / "data" / "generated" / "base-game.json"
HANDLERS = json.loads((Path(__file__).parent / "effect-handlers.json").read_text(encoding="utf-8"))
REQUIRED_SHEETS = [
    "README", "Actions", "Marines", "Events", "Terrain", "Locations",
    "Location Terrain", "Genestealers", "Brood Lords", "Setup & Rules",
]
DATA_VERSION = "base-game-v2"


def clean(value):
    return None if value is None else str(value).strip()


def integer(value):
    result = int(value)
    assert result == float(value), f"Expected integer, received {value}"
    return result


def truthy(value):
    return value is True or value == 1 or str(value).lower() == "true"


def slug(value):
    normalized = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode()
    normalized = normalized.replace("'", "").replace("&", " and ").lower()
    return re.sub(r"^-|-$", "", re.sub(r"[^a-z0-9]+", "-", normalized))


def source(sheet, row):
    return {"workbook": DATA_VERSION, "sheet": sheet, "row": row}


workbook = load_workbook(INPUT, data_only=True, read_only=True)
assert workbook.sheetnames == REQUIRED_SHEETS, workbook.sheetnames


def table_rows(sheet_name):
    sheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in sheet[1]]
    result = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells]
        if not any(value is not None and str(value).strip() for value in values):
            continue
        result.append((row_number, dict(zip(headers, values))))
    return result


def handler(group, name):
    result = HANDLERS[group].get(name)
    assert result, f"Missing {group} effect handler for {name}"
    return result


actions = []
for row, value in table_rows("Actions"):
    actions.append({
        "id": f"action.{slug(value['Team'])}.{slug(value['Action'])}",
        "team": clean(value["Team"]).upper(),
        "name": clean(value["Action"]),
        "initiative": integer(value["Initiative"]),
        "type": clean(value["Type"]).upper().replace(" + ", "_").replace(" ", "_"),
        "sourceText": clean(value["Exact Gameplay Text"]),
        "target": clean(value["Named/Relevant Target"]),
        "timing": clean(value["Timing"]),
        "handlerId": handler("actions", clean(value["Action"])),
        "source": source("Actions", row),
    })

marines = []
for row, value in table_rows("Marines"):
    assert clean(value["Facing"]) == clean(value["Team"]), f"Unexpected Marines!E{row}"
    marines.append({
        "id": f"marine.{slug(value['Team'])}.{slug(value['Marine'])}",
        "team": clean(value["Team"]).upper(),
        "name": clean(value["Marine"]),
        "attackRange": integer(value["Attack Range"]),
        "namedActionAbility": clean(value["Named Action Ability"]),
        "source": source("Marines", row),
    })

events = []
for row, value in table_rows("Events"):
    source_name = clean(value["Event"])
    copy_match = re.fullmatch(r"(.+?) \(Copy (\d+)\)", source_name)
    name = copy_match.group(1) if copy_match else source_name
    copy_index = integer(copy_match.group(2)) if copy_match else None
    event_id = f"event.{slug(name)}" + (f".copy-{copy_index}" if copy_index else "")
    events.append({
        "id": event_id,
        "name": name,
        "copyIndex": copy_index,
        "quantity": integer(value["Qty"]),
        "instinct": truthy(value["Instinct"]),
        "sourceText": clean(value["Exact Gameplay Text"]),
        "activations": [
            {
                "severity": clean(value[f"Activation {index} Severity"]).upper(),
                "terrainColor": clean(value[f"Activation {index} Terrain Color"]).upper(),
            }
            for index in (1, 2)
        ],
        "movement": clean(value["Movement"]).upper() if clean(value["Movement"]) else None,
        "movementIcon": clean(value["Movement Icon"]).upper() if clean(value["Movement Icon"]) else None,
        "handlerId": handler("events", name),
        "source": source("Events", row),
    })

event_variants = defaultdict(list)
for event in events:
    if event["copyIndex"] is not None:
        event_variants[event["name"]].append(event)
for name, variants in event_variants.items():
    variants.sort(key=lambda item: item["copyIndex"])
    assert [item["copyIndex"] for item in variants] == list(range(1, len(variants) + 1)), f"Invalid copy numbering for {name}"
    assert all(item["quantity"] == 1 for item in variants), f"Event variants must be physical quantity 1: {name}"
    assert len({item["movementIcon"] for item in variants}) == len(variants), f"Event variants must have distinct movement icons: {name}"
    shared_fields = ("instinct", "sourceText", "activations", "movement", "handlerId")
    assert all(tuple(item[field] for field in shared_fields) == tuple(variants[0][field] for field in shared_fields) for item in variants[1:]), f"Unexpected non-icon variant difference for {name}"

terrain = []
for row, value in table_rows("Terrain"):
    name = clean(value["Terrain"])
    terrain.append({
        "id": f"terrain.{slug(name)}",
        "name": name,
        "spawnColor": clean(value["Spawn Color"]).upper(),
        "activatable": truthy(value["Activatable"]),
        "sourceText": clean(value["Exact Gameplay Text"]),
        "handlerId": handler("terrain", name),
        "source": source("Terrain", row),
    })

locations = []
for row, value in table_rows("Locations"):
    name = clean(value["Location"])
    locations.append({
        "id": f"location.{slug(name)}",
        "name": name,
        "tier": clean(value["Deck Tier"]),
        "leftBlips": integer(value["Left Blips"]),
        "rightBlips": integer(value["Right Blips"]),
        "abilityTiming": clean(value["Ability Timing"]),
        "sourceText": clean(value["Exact Gameplay Text"]),
        "handlerId": handler("locations", name),
        "source": source("Locations", row),
    })

terrain_by_name = {item["name"]: item["id"] for item in terrain}
location_by_name = {item["name"]: item["id"] for item in locations}
location_terrain = []
for row, value in table_rows("Location Terrain"):
    location_terrain.append({
        "locationId": location_by_name.get(clean(value["Location"])),
        "side": clean(value["Side"]).upper(),
        "markerOrder": integer(value["Marker Order"]),
        "terrainId": terrain_by_name.get(clean(value["Terrain"])),
        "distance": integer(value["Distance"]),
        "countFrom": clean(value["Count From"]).upper(),
        "source": source("Location Terrain", row),
    })
assert all(item["locationId"] and item["terrainId"] for item in location_terrain)

genestealer_types = []
for row, value in table_rows("Genestealers"):
    genestealer_types.append({
        "id": f"genestealer.{slug(value['Icon'])}",
        "icon": clean(value["Icon"]).upper(),
        "quantity": integer(value["Qty"]),
        "source": source("Genestealers", row),
    })

brood_lords = []
for row, value in table_rows("Brood Lords"):
    brood_lords.append({
        "id": f"brood-lord.{slug(clean(value['Card']).replace('Brood Lord ', ''))}",
        "name": clean(value["Card"]),
        "quantity": integer(value["Qty"]),
        "movementIcons": [part.strip().upper() for part in clean(value["Movement Icons"]).split("+")],
        "sourceText": clean(value["Gameplay Rules"]),
        "source": source("Brood Lords", row),
    })

setup_sheet = workbook["Setup & Rules"]
player_setups = []
for row in range(2, 8):
    player_setups.append({
        "players": integer(re.search(r"\d+", clean(setup_sheet.cell(row, 1).value)).group()),
        "formationSize": integer(re.search(r"\d+", clean(setup_sheet.cell(row, 2).value)).group()),
        "combatTeams": clean(setup_sheet.cell(row, 3).value),
        "locationDeckSetup": [part.strip() for part in clean(setup_sheet.cell(row, 4).value).split(">")],
        "majorSpawn": integer(setup_sheet.cell(row, 5).value),
        "minorSpawn": integer(setup_sheet.cell(row, 6).value),
        "startLeftBlips": integer(setup_sheet.cell(row, 7).value),
        "startRightBlips": integer(setup_sheet.cell(row, 8).value),
        "source": source("Setup & Rules", row),
    })

setup_terrain = []
for row in range(17, 33):
    setup_terrain.append({
        "setup": clean(setup_sheet.cell(row, 1).value),
        "side": clean(setup_sheet.cell(row, 2).value).upper(),
        "markerOrder": integer(setup_sheet.cell(row, 3).value),
        "terrainId": terrain_by_name[clean(setup_sheet.cell(row, 4).value)],
        "distance": integer(setup_sheet.cell(row, 5).value),
        "countFrom": clean(setup_sheet.cell(row, 6).value).upper(),
        "source": source("Setup & Rules", row),
    })

setup_location_groups = {
    "Void Lock - 1 player": [1],
    "Void Lock - 2 or 4 players": [2, 4],
    "Void Lock - 5 players": [5],
    "Void Lock - 3 or 6 players": [3, 6],
}
setup_locations = []
for setup_name, player_counts in setup_location_groups.items():
    setup_locations.append({
        "id": f"setup-location.{slug(setup_name)}",
        "name": setup_name,
        "playerCounts": player_counts,
        "source": source("Setup & Rules", player_counts[0] + 1),
    })

setup_location_by_player = {
    players: item["id"]
    for item in setup_locations
    for players in item["playerCounts"]
}
for item in player_setups:
    item["setupLocationId"] = setup_location_by_player[item["players"]]

component_constants = {
    slug(setup_sheet.cell(row, 14).value): integer(setup_sheet.cell(row, 15).value)
    for row in range(2, 10)
}


def copies(definitions):
    return [
        {"id": f"{item['id']}.{index:02d}", "definitionId": item["id"]}
        for item in definitions
        for index in range(1, item["quantity"] + 1)
    ]


instances = {
    "actions": [{"id": item["id"], "definitionId": item["id"]} for item in actions],
    "marines": [{"id": item["id"], "definitionId": item["id"]} for item in marines],
    "events": copies(events),
    "genestealers": copies(genestealer_types),
    "broodLords": copies(brood_lords),
    "terrain": [{"id": item["id"], "definitionId": item["id"]} for item in terrain],
    "locations": [{"id": item["id"], "definitionId": item["id"]} for item in locations],
    "setupLocations": [
        {"id": item["id"], "definitionId": item["id"]}
        for item in setup_locations
    ],
}

assert len(actions) == 18
assert sorted(item["initiative"] for item in actions) == list(range(1, 19))
assert set(Counter(item["team"] for item in actions).values()) == {3}
assert len(marines) == 12 and set(Counter(item["team"] for item in marines).values()) == {2}
assert len(instances["events"]) == 30
assert len(events) == 24
assert len(terrain) == 8
assert len(locations) + 4 == 22
assert len(instances["terrain"]) == 8
assert len(instances["locations"]) == 18
assert len(instances["setupLocations"]) == 4
assert len(location_terrain) == 72
placement_counts = Counter(item["locationId"] for item in location_terrain)
assert len(placement_counts) == 18 and set(placement_counts.values()) == {4}
for location_id in placement_counts:
    terrain_ids = [item["terrainId"] for item in location_terrain if item["locationId"] == location_id]
    assert len(terrain_ids) == len(set(terrain_ids)), f"Duplicate Terrain type on {location_id}"
assert len(instances["genestealers"]) == 36
assert len(instances["broodLords"]) == 2

generated = {
    "schemaVersion": "1.0.0",
    "dataVersion": DATA_VERSION,
    "generatedFrom": INPUT.name,
    "definitions": {
        "actions": actions,
        "marines": marines,
        "events": events,
        "terrain": terrain,
        "locations": locations,
        "setupLocations": setup_locations,
        "locationTerrain": location_terrain,
        "genestealerTypes": genestealer_types,
        "broodLords": brood_lords,
    },
    "setup": {
        "playerSetups": player_setups,
        "setupTerrain": setup_terrain,
        "componentConstants": component_constants,
    },
    "instances": instances,
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(generated, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
print(f"Generated {OUTPUT}")
print("18 actions, 12 marines, 30 events, 22 locations, 8 terrain, 36 genestealers, 2 brood lords")
